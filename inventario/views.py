from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import transaction
from django.db.models import Q, Sum, Count, Max
from django.utils import timezone
from django.http import JsonResponse
from datetime import timedelta
from django.db.models.functions import TruncMonth
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth.models import Group
from .forms import HerramientaForm
from django.http import JsonResponse
from .recomendador import recomendar_herramientas
from . import recomendador as rec
from datetime import datetime
from django.core.paginator import Paginator
#
from django.http import HttpResponse
import io
import openpyxl
from openpyxl.utils import get_column_letter

from xhtml2pdf import pisa

from .models import (
    Herramienta,
    Docente,
    Estudiante,
    Panolero,
    Asignatura,
    Prestamo,
    PrestamoDetalle,
    Preparacion,
    PreparacionDetalle,
    Baja,
    BajaDetalle
)

## Nuevo
from .forms import HerramientaForm, CrearUsuarioForm


def es_jefe_panol(user):
    if not user.is_authenticated:
        return False

    pan = Panolero.objects.filter(user=user, activo=True).first()
    if not pan:
        return False

    # En tu tabla, el rol de jefatura es 'jefe'
    return (pan.rol or "").lower() == "jefe"




# ---------------------------------------------
# VISTA: ADMINISTRACIÓN DE USUARIOS
# ---------------------------------------------

@login_required
@user_passes_test(es_jefe_panol)
def administracion(request):
    if request.method == "POST":
        form = CrearUsuarioForm(request.POST)
        if form.is_valid():
            user = None
            try:
                # 1) Crear usuario Django (auth_user)
                user = form.save(commit=False)
                user.is_staff = False      # no damos staff por defecto
                user.is_superuser = False  # ni superusuario
                user.save()

                # 2) Datos de apoyo
                rol = (form.cleaned_data.get("rol") or "").upper().strip()
                nombre_completo = f"{user.first_name} {user.last_name}".strip()

                # 3) Si es PAÑOLERO
                if rol == "PANOLERO":
                    grupo, _ = Group.objects.get_or_create(name="Pañolero")
                    user.groups.add(grupo)

                    Panolero.objects.create(
                        user=user,
                        nombre=nombre_completo,
                        rol="panolero",  # texto que usas en la tabla
                        activo=True,
                    )

                # 4) Si es DOCENTE
                elif rol == "DOCENTE":
                    grupo, _ = Group.objects.get_or_create(name="Docente")
                    user.groups.add(grupo)

                    # Buscamos el último código numérico y sumamos 1
                    ultimo_codigo = Docente.objects.aggregate(
                        max_codigo=Max("codigo")
                    )["max_codigo"] or 0
                    nuevo_codigo = ultimo_codigo + 1

                    Docente.objects.create(
                        codigo=nuevo_codigo,      # ej: 40001, 40002, ...
                        nombre=nombre_completo,   # ej: "JUAN PÉREZ"
                        activo=True,
                    )

                else:
                    messages.error(request, f"Rol no reconocido: {rol}")
                    if user:
                        user.delete()
                    return render(
                        request,
                        "inventario/administracion.html",
                        {"form": form},
                    )

                messages.success(request, "Usuario creado correctamente.")
                return redirect("administracion")

            except Exception as e:
                # Si algo falla al crear Panolero/Docente, borramos el user Django
                if user:
                    try:
                        user.delete()
                    except Exception:
                        pass
                messages.error(
                    request,
                    f"Ocurrió un error al crear el usuario: {e}",
                )
        else:
            # Formulario no válido: errores en username, contraseña, etc.
            messages.error(request, "Revisa los datos del formulario.")
    else:
        form = CrearUsuarioForm()

    return render(request, "inventario/administracion.html", {"form": form})




# ---------------------------------------------------
# 1) LISTA DE HERRAMIENTAS (INVENTARIO)
# ---------------------------------------------------
@login_required
def lista_herramientas(request):
    query = request.GET.get("q", "")

    if query:
        herramientas = Herramienta.objects.filter(
            Q(nombre__icontains=query) | Q(codigo__icontains=query)
        )
    else:
        herramientas = Herramienta.objects.all()

    return render(request, "inventario/lista_herramientas.html", {
        "herramientas": herramientas,
        "query": query,
    })


#api de herramienta
@login_required
def api_herramienta_por_codigo(request):
    """
    Devuelve info básica de la herramienta dado su código o código de barra, en formato JSON.
    Esto lo usa el escáner/página de crear_prestamo y crear_preparacion.
    """
    codigo = request.GET.get("codigo", "").strip()

    if not codigo:
        return JsonResponse(
            {"ok": False, "error": "Código vacío."},
            status=400
        )

    # 1) Buscamos por 'codigo'
    h = Herramienta.objects.filter(codigo=codigo).first()

    # 2) Si no está, probamos por 'codigo_barra'
    if h is None:
        h = Herramienta.objects.filter(codigo_barra=codigo).first()

    if h is None:
        return JsonResponse(
            {"ok": False, "error": "Herramienta no encontrada."},
            status=404
        )

    return JsonResponse({
        "ok": True,
        "codigo": h.codigo,
        "codigo_barra": h.codigo_barra,
        "nombre": h.nombre,
        "stock_disponible": h.stock_disponible,
    })



# ---------------------------------------------------
# HELPERS PARA ROLES
# ---------------------------------------------------
def es_panolero(user):
    return (
        user.is_authenticated
        and user.groups.filter(name="Pañolero").exists()
    )


def es_jefe_panol(user):
    return (
        user.is_authenticated
        and user.groups.filter(name="JefePañol").exists()
    )


def obtener_panolero_desde_user(user):
    """Devuelve el objeto Panolero asociado al usuario logueado."""
    if not user.is_authenticated:
        return None
    return Panolero.objects.filter(user=user, activo=True).first()


# ---------------------------------------------------
# 2) GESTIONAR HERRAMIENTA (CREAR + SUMAR STOCK)
# ---------------------------------------------------
@login_required
def gestionar_herramienta(request):
    mensaje = None
    error = None

    # Inventario completo para el listado de abajo
    herramientas = Herramienta.objects.all().order_by("codigo")

    # ---------- BUSCADOR PARA SUMAR STOCK (GET: q_stock) ----------
    q_stock = request.GET.get("q_stock", "").strip()

    if q_stock:
        herramientas_busqueda = Herramienta.objects.filter(
            Q(nombre__icontains=q_stock) |
            Q(codigo__icontains=q_stock)
        ).order_by("codigo")
    else:
        # Si no se está filtrando, mostramos todas en el combo
        herramientas_busqueda = herramientas

    # ---------- ACCIONES POR POST (CREAR / SUMAR STOCK) ----------
    if request.method == "POST":
        accion = request.POST.get("accion")

        # 1) CREAR NUEVA HERRAMIENTA
        if accion == "crear":
            nombre = request.POST.get("nombre", "").strip()
            tipo = request.POST.get("tipo", "").strip() or "herramienta"
            stock_inicial = request.POST.get("stock", "0").strip()

            try:
                stock_valor = int(stock_inicial)
                if stock_valor < 0:
                    stock_valor = 0
            except ValueError:
                stock_valor = 0

            if not nombre:
                error = "El nombre de la herramienta es obligatorio."
            else:
                h = Herramienta(
                    nombre=nombre,
                    tipo=tipo,
                    stock=stock_valor,
                    stock_disponible=stock_valor,
                )
                # save() autogenera código/código_barra
                h.save()
                mensaje = (
                    f"Herramienta '{h.nombre}' creada con éxito. "
                    f"Código asignado: {h.codigo}."
                )

        # 2) SUMAR STOCK A UNA HERRAMIENTA EXISTENTE
        elif accion == "sumar_stock":
            codigo = request.POST.get("codigo_existente", "").strip()
            cantidad = request.POST.get("cantidad_sumar", "0").strip()

            try:
                cantidad_valor = int(cantidad)
            except ValueError:
                cantidad_valor = 0

            if not codigo or cantidad_valor <= 0:
                error = "Debes seleccionar una herramienta y una cantidad válida."
            else:
                herramienta = get_object_or_404(Herramienta, codigo=codigo)

                # 🔒 NO permitir sumar stock a llaves (llave / llave_auto)
                tipo_h = (herramienta.tipo or "").strip().lower()
                if tipo_h.startswith("llave"):
                    error = (
                        f"No se permite sumar stock a la herramienta '{herramienta.nombre}' "
                        f"porque está marcada como llave."
                    )
                else:
                    herramienta.stock += cantidad_valor
                    herramienta.stock_disponible += cantidad_valor
                    herramienta.save()
                    mensaje = (
                        f"Se sumaron {cantidad_valor} unidades a '{herramienta.nombre}'. "
                        f"Stock actual: {herramienta.stock}."
                    )

    context = {
        "herramientas": herramientas,              # para tabla de inventario
        "herramientas_busqueda": herramientas_busqueda,  # para el <select>
        "q_stock": q_stock,                        # para mantener el texto buscado
        "mensaje": mensaje,
        "error": error,
    }

    return render(request, "inventario/gestionar_herramienta.html", context)


# ---------------------------------------------------
# 3) LISTA DE PRÉSTAMOS
# ---------------------------------------------------
@login_required
def lista_prestamos(request):
    q = request.GET.get("q", "").strip()

    prestamos = Prestamo.objects.select_related(
        "panolero",
        "docente",
        "estudiante",   
        "asignatura",
    ).all()

    if q:
        prestamos = prestamos.filter(
            Q(codigo_prestamo__icontains=q)
            | Q(docente__nombre__icontains=q)
            | Q(estudiante__nombre__icontains=q)   # 👈 opcional: buscar por alumno
            | Q(panolero__nombre__icontains=q)
            | Q(asignatura__nombre__icontains=q)
        )

    prestamos = prestamos.order_by("-fecha", "-id")

    return render(request, "inventario/lista_prestamos.html", {
        "prestamos": prestamos,
        "query": q,
    })



# ---------------------------------------------------
# 4) CREAR PRÉSTAMO (REGISTRAR SALIDA)
#   Formulario tipo planilla VACÍO.
#   El JS usa 'herramientas' sólo para armar el diccionario INVENTARIO.
# ---------------------------------------------------
@login_required
def crear_prestamo(request):
    mensaje = None
    error = None

    # 👉 Sólo pañolero puede registrar préstamos
    panolero = obtener_panolero_desde_user(request.user)
    if panolero is None:
        error = (
            "Tu usuario no está asociado a ningún pañolero activo. "
            "Pide a la jefatura que te registre en el módulo de pañoleros."
        )
        return render(request, "inventario/crear_prestamo.html", {
            "mensaje": mensaje,
            "error": error,
            "docentes": Docente.objects.filter(activo=True).order_by("nombre"),
            "asignaturas": Asignatura.objects.all().order_by("nombre"),
            "herramientas": Herramienta.objects.all().order_by("nombre"),
        })

    if request.method == "POST":
        tipo_solicitante   = request.POST.get("tipo_solicitante", "docente")
        docente_codigo     = request.POST.get("docente_codigo", "").strip()

        # Datos estudiante
        est_rut            = request.POST.get("estudiante_rut", "").strip()
        est_nombre         = request.POST.get("estudiante_nombre", "").strip()
        est_carrera        = request.POST.get("estudiante_carrera", "").strip()

        # Asignatura (texto, y opcionalmente id si algún día usas <select>)
        asignatura_id      = request.POST.get("asignatura_id", "").strip()
        asignatura_nombre  = request.POST.get("asignatura_nombre", "").strip()

        fecha_str          = request.POST.get("fecha", "").strip()
        hora_inicio_str    = request.POST.get("hora_inicio", "").strip()
        hora_fin_str       = request.POST.get("hora_fin", "").strip()
        observaciones      = request.POST.get("observaciones", "").strip()

        # Código de preparación (hidden)
        codigo_preparacion_origen = request.POST.get("codigo_preparacion_origen", "").strip()
        prep_origen = None
        if codigo_preparacion_origen:
            prep_origen = Preparacion.objects.filter(
                codigo_preparacion=codigo_preparacion_origen
            ).first()

        # Si la preparación existe pero ya está usada/anulada, no permitir
        if prep_origen and prep_origen.estado in ["usado", "anulado"]:
            error = (
                f"La preparación {prep_origen.codigo_preparacion} ya fue "
                f"marcada como entregada/anulada y no puede volver a generar un préstamo."
            )
            return render(request, "inventario/crear_prestamo.html", {
                "mensaje": None,
                "error": error,
                "docentes": Docente.objects.filter(activo=True).order_by("nombre"),
                "asignaturas": Asignatura.objects.all().order_by("nombre"),
                "herramientas": Herramienta.objects.all().order_by("nombre"),
            })

        # Si el tipo es estudiante, no debería usar preparación anticipada
        if prep_origen and tipo_solicitante == "estudiante":
            error = "Las preparaciones anticipadas solo pueden usarse con préstamos de docentes."
            return render(request, "inventario/crear_prestamo.html", {
                "mensaje": None,
                "error": error,
                "docentes": Docente.objects.filter(activo=True).order_by("nombre"),
                "asignaturas": Asignatura.objects.all().order_by("nombre"),
                "herramientas": Herramienta.objects.all().order_by("nombre"),
            })

        codigos    = request.POST.getlist("codigo_herramienta")
        cantidades = request.POST.getlist("cantidad")

        docente    = None
        estudiante = None
        asignatura = None

        # ---------- DOCENTE / ESTUDIANTE ----------
        if tipo_solicitante == "docente":
            if docente_codigo:
                docente = Docente.objects.filter(
                    codigo=docente_codigo,
                    activo=True
                ).first()
                if docente is None:
                    error = "Docente no válido o inactivo."
            elif prep_origen and prep_origen.docente:
                docente = prep_origen.docente
            else:
                error = "Debes seleccionar un docente."

        elif tipo_solicitante == "estudiante":
            if not est_rut or not est_nombre:
                error = "Debes indicar RUT y nombre del estudiante."
            else:
                estudiante, _ = Estudiante.objects.get_or_create(
                    rut=est_rut,
                    defaults={
                        "nombre": est_nombre,
                        "carrera": est_carrera,
                        "activo": True,
                    },
                )

        # ---------- ASIGNATURA ----------
        if tipo_solicitante == "docente":
            if asignatura_id:
                asignatura = Asignatura.objects.filter(id=asignatura_id).first()
            elif asignatura_nombre:
                asignatura, _ = Asignatura.objects.get_or_create(
                    nombre=asignatura_nombre,
                    defaults={"codigo": None},
                )
            elif prep_origen and prep_origen.asignatura:
                asignatura = prep_origen.asignatura

            if asignatura is None and not error:
                error = "Debes indicar la asignatura (o que la preparación la tenga)."
        else:
            # Para estudiantes la asignatura es opcional
            if asignatura_nombre:
                asignatura, _ = Asignatura.objects.get_or_create(
                    nombre=asignatura_nombre,
                    defaults={"codigo": None},
                )

        # ---------- FECHA ----------
        try:
            fecha = timezone.datetime.strptime(fecha_str, "%Y-%m-%d").date()
        except Exception:
            if prep_origen and prep_origen.fecha:
                fecha = prep_origen.fecha
            else:
                fecha = timezone.now().date()

        # ---------- HORAS ----------
        try:
            hora_inicio = timezone.datetime.strptime(hora_inicio_str, "%H:%M").time()
        except Exception:
            hora_inicio = None

        try:
            hora_fin = timezone.datetime.strptime(hora_fin_str, "%H:%M").time()
        except Exception:
            hora_fin = None

        if hora_inicio is None and prep_origen and prep_origen.hora_inicio:
            hora_inicio = prep_origen.hora_inicio
        if hora_fin is None and prep_origen and prep_origen.hora_fin:
            hora_fin = prep_origen.hora_fin

        if hora_inicio is None and not error:
            error = "Debes indicar una hora de inicio (en el formulario o en la preparación)."

        # ---------- LÍNEAS DETALLE ----------
        lineas_validas = []
        for c, cant in zip(codigos, cantidades):
            c    = c.strip()
            cant = cant.strip()
            if not c or not cant:
                continue
            try:
                cant_int = int(cant)
            except ValueError:
                continue
            if cant_int <= 0:
                continue
            lineas_validas.append((c, cant_int))

        if not lineas_validas and not error:
            error = "Debes ingresar al menos una herramienta con cantidad mayor a 0."

        if error:
            return render(request, "inventario/crear_prestamo.html", {
                "mensaje": None,
                "error": error,
                "docentes": Docente.objects.filter(activo=True).order_by("nombre"),
                "asignaturas": Asignatura.objects.all().order_by("nombre"),
                "herramientas": Herramienta.objects.all().order_by("nombre"),
            })

        # ---------------------------------------------
        # REGISTRO DEL PRÉSTAMO
        # ---------------------------------------------
        try:
            with transaction.atomic():
                codigo_prestamo = "P" + timezone.now().strftime("%Y%m%d%H%M%S")

                prestamo = Prestamo.objects.create(
                    codigo_prestamo=codigo_prestamo,
                    fecha=fecha,
                    hora_inicio=hora_inicio,
                    hora_fin=hora_fin,
                    panolero=panolero,
                    docente=docente if tipo_solicitante == "docente" else None,
                    estudiante=estudiante if tipo_solicitante == "estudiante" else None,
                    asignatura=asignatura,
                    estado="pendiente",
                    observaciones=observaciones or (
                        prep_origen.observaciones if prep_origen else ""
                    ),
                )

                # Flag para ver si TODAS las líneas son consumibles
                solo_consumibles  = True
                # True si viene de una preparación (el stock_disponible NO se tocó en la preparación)
                desde_preparacion = prep_origen is not None

                for codigo, cantidad in lineas_validas:
                    # Buscar por código o código de barra
                    herramienta = Herramienta.objects.filter(codigo=codigo).first()
                    if herramienta is None:
                        herramienta = Herramienta.objects.filter(
                            codigo_barra=codigo
                        ).first()

                    if herramienta is None:
                        raise ValueError(
                            f"No se encontró ninguna herramienta con código/código de barra '{codigo}'."
                        )

                    tipo_herr = (herramienta.tipo or "").strip().lower()
                    # 👇 más tolerante: cualquier tipo que contenga "consum"
                    es_consumible = "consum" in tipo_herr

                    # 🚫 Estudiante no puede pedir llaves / llaves de auto
                    if tipo_solicitante == "estudiante" and tipo_herr.startswith("llave"):
                        raise ValueError(
                            f"Las llaves solo pueden ser prestadas a docentes. "
                            f"Herramienta: {herramienta.nombre}"
                        )

                    # -------- VALIDACIÓN DE STOCK --------
                    if desde_preparacion:
                        # Viene de una preparación: usamos stock físico disponible,
                        # la planificación ya consideró disponibilidad.
                        if herramienta.stock_disponible < cantidad:
                            raise ValueError(
                                f"No hay suficiente stock disponible para {herramienta.nombre} "
                                f"al momento de entregar la preparación. "
                                f"Disponible: {herramienta.stock_disponible}, solicitado: {cantidad}"
                            )
                    else:
                        # Préstamo directo: respetar preparaciones en los próximos 15 minutos
                        stock_efectivo = stock_disponible_respetando_preps(herramienta)
                        if stock_efectivo < cantidad:
                            raise ValueError(
                                f"No hay suficiente stock disponible para {herramienta.nombre} "
                                f"considerando preparaciones próximas. "
                                f"Disponible efectivo: {max(stock_efectivo, 0)}, "
                                f"solicitado: {cantidad}"
                            )

                    # Descontar SIEMPRE del stock_disponible al concretar el préstamo
                    herramienta.stock_disponible -= cantidad
                    if herramienta.stock_disponible < 0:
                        herramienta.stock_disponible = 0

                    # Ajuste de stock TOTAL:
                    # - Consumible: se consume al prestar (venga o no de preparación)
                    # - No consumible: stock total no cambia, solo disponible.
                    if es_consumible:
                        herramienta.stock -= cantidad
                        if herramienta.stock < 0:
                            herramienta.stock = 0
                    else:
                        solo_consumibles = False

                    herramienta.save()

                    # Crear detalle de préstamo
                    PrestamoDetalle.objects.create(
                        prestamo=prestamo,
                        herramienta=herramienta,
                        cantidad_solicitada=cantidad,
                        cantidad_entregada=cantidad,
                        cantidad_devuelta=0
                    )

                # ⬇⬇⬇ CIERRE AUTOMÁTICO SI TODO ES CONSUMIBLE
                if solo_consumibles:
                    prestamo.estado = "devuelto"
                    prestamo.save(update_fields=["estado"])
                # Si hay mezcla o solo no-consumibles, queda "pendiente"

                if prep_origen is not None:
                    prep_origen.estado = "usado"
                    prep_origen.save(update_fields=["estado", "updated_at"])

                mensaje = f"Préstamo creado correctamente. Código: {prestamo.codigo_prestamo}"

        except Exception as e:
            error = f"Error: {e}"

        return render(request, "inventario/crear_prestamo.html", {
            "mensaje": mensaje,
            "error": error,
            "docentes": Docente.objects.filter(activo=True).order_by("nombre"),
            "asignaturas": Asignatura.objects.all().order_by("nombre"),
            "herramientas": Herramienta.objects.all().order_by("nombre"),
        })

    # GET → formulario vacío
    return render(request, "inventario/crear_prestamo.html", {
        "mensaje": None,
        "error": None,
        "docentes": Docente.objects.filter(activo=True).order_by("nombre"),
        "asignaturas": Asignatura.objects.all().order_by("nombre"),
        "herramientas": Herramienta.objects.all().order_by("nombre"),
    })


# ---------------------------------------------------
# 5) REGISTRAR DEVOLUCIÓN
# ---------------------------------------------------
@login_required
def registrar_devolucion(request, prestamo_id):
    """
    Registra la devolución parcial o total de un préstamo.
    - Ajusta stock_disponible de las herramientas.
    - Si la herramienta es consumible, también ajusta el stock total.
    - Actualiza el estado del préstamo:
      * Si todas las herramientas NO consumibles están devueltas → devuelto.
      * Si falta alguna NO consumible → devuelto_parcial.
      * Los consumibles no devueltos no bloquean el cierre del préstamo.
    """
    prestamo = get_object_or_404(
        Prestamo.objects.prefetch_related("detalles__herramienta"),
        id=prestamo_id
    )

    mensaje = None
    error = None

    # Solo lectura si el préstamo está anulado
    solo_lectura = prestamo.estado in ["anulado"]

    if request.method == "POST" and not solo_lectura:
        try:
            with transaction.atomic():
                for detalle in prestamo.detalles.all():
                    campo = f"detalle_{detalle.id}_devuelta"
                    valor_str = request.POST.get(campo, "").strip()

                    # Si no viene nada, mantenemos la cantidad actual
                    try:
                        nueva_cantidad = int(valor_str)
                    except (ValueError, TypeError):
                        nueva_cantidad = detalle.cantidad_devuelta

                    # No permitir negativos
                    if nueva_cantidad < 0:
                        nueva_cantidad = 0

                    # No permitir devolver más de lo entregado
                    if nueva_cantidad > detalle.cantidad_entregada:
                        nueva_cantidad = detalle.cantidad_entregada

                    # Diferencia vs lo que ya estaba devuelto
                    delta = nueva_cantidad - detalle.cantidad_devuelta

                    if delta != 0:
                        detalle.cantidad_devuelta = nueva_cantidad
                        detalle.save()

                        herramienta = detalle.herramienta

                        # Siempre ajustamos el stock disponible
                        herramienta.stock_disponible += delta

                        # Si es consumible, también ajustamos el stock total
                        tipo_h = (herramienta.tipo or "").strip().lower()
                        es_consumible = "consum" in tipo_h

                        if es_consumible:
                            herramienta.stock += delta
                            if herramienta.stock < 0:
                                herramienta.stock = 0

                        if herramienta.stock_disponible < 0:
                            herramienta.stock_disponible = 0

                        herramienta.save()

                # ----------- NUEVA LÓGICA DE ESTADO -----------
                detalles = list(prestamo.detalles.all())

                # Consideramos solo herramientas NO consumibles
                detalles_no_consumibles = []
                for d in detalles:
                    tipo_h = (d.herramienta.tipo or "").strip().lower()
                    es_consumible = "consum" in tipo_h
                    if not es_consumible:
                        detalles_no_consumibles.append(d)

                if not detalles_no_consumibles:
                    # Si no hay no-consumibles, todo es consumible → lo consideramos devuelto
                    prestamo.estado = "devuelto"
                else:
                    # El estado depende SOLO de las herramientas no consumibles
                    todos_no_consumibles_completos = all(
                        d.cantidad_devuelta == d.cantidad_entregada
                        for d in detalles_no_consumibles
                    )

                    if todos_no_consumibles_completos:
                        prestamo.estado = "devuelto"
                    else:
                        prestamo.estado = "devuelto_parcial"
                # ---------------------------------------------

                bitacora_texto = request.POST.get("bitacora_devolucion", "").strip()
                prestamo.bitacora_devolucion = bitacora_texto

                prestamo.save()
                mensaje = "Devolución registrada correctamente."
                solo_lectura = prestamo.estado in ["anulado"]

        except Exception as e:
            error = f"Ocurrió un error al registrar la devolución: {e}"

    return render(request, "inventario/registrar_devolucion.html", {
        "prestamo": prestamo,
        "mensaje": mensaje,
        "error": error,
        "solo_lectura": solo_lectura,
    })

# ---------------------------------------------------
# 6) LISTA DE PREPARACIONES (LISTADO DE CLASES)
# ---------------------------------------------------
@login_required
def lista_preparaciones(request):
    q = request.GET.get("q", "").strip()
    fecha_str = request.GET.get("fecha", "").strip()

    preparaciones = Preparacion.objects.select_related(
        "panolero", "docente", "asignatura"
    ).all()

    # Filtrar por fecha, si viene
    if fecha_str:
        try:
            fecha = timezone.datetime.strptime(fecha_str, "%Y-%m-%d").date()
            preparaciones = preparaciones.filter(fecha=fecha)
        except ValueError:
            pass

    # Búsqueda general
    if q:
        preparaciones = preparaciones.filter(
            Q(codigo_preparacion__icontains=q)
            | Q(docente__nombre__icontains=q)
            | Q(panolero__nombre__icontains=q)
            | Q(asignatura__nombre__icontains=q)
            | Q(estado__icontains=q)
        )

    preparaciones = preparaciones.order_by("-fecha", "-id")

    return render(request, "inventario/lista_preparaciones.html", {
        "preparaciones": preparaciones,
        "query": q,
        "fecha": fecha_str,
    })





# ---------------------------------------------------
# 7) CREAR PREPARACIÓN (LISTADO DE CLASE / PICKING)
# ---------------------------------------------------
@login_required
def crear_preparacion(request):
    mensaje = None
    error = None

    user = request.user

    # 👉 ¿Este usuario pertenece al grupo "Docente"?
    es_docente = user.groups.filter(name="Docente").exists()

    # 👉 Intentamos vincularlo a un registro de la tabla Docente por nombre completo
    docente_actual = None
    if es_docente:
        nombre_full = (f"{user.first_name} {user.last_name}").strip()
        if nombre_full:
            docente_actual = Docente.objects.filter(
                nombre__iexact=nombre_full,
                activo=True
            ).first()

    # 👉 Pañolero asociado al usuario (para perfil pañolero)
    panolero = obtener_panolero_desde_user(user)

    # 👉 Si no hay pañolero y es DOCENTE,
    #    usamos por defecto la JEFATURA como responsable de la preparación
    if panolero is None and es_docente:
        panolero = Panolero.objects.filter(
            rol__in=["jefe", "Jefe", "jefatura", "Jefatura"],
            activo=True
        ).first()

    # 👉 Si NO es docente y NO tenemos pañolero, bloqueamos
    if panolero is None and not es_docente:
        error = (
            "Tu usuario no está asociado a ningún pañolero activo. "
            "Pide a la jefatura que te registre en el módulo de pañoleros."
        )
        return render(request, "inventario/crear_preparacion.html", {
            "mensaje": mensaje,
            "error": error,
            "docentes": Docente.objects.filter(activo=True).order_by("nombre"),
            "asignaturas": Asignatura.objects.all().order_by("nombre"),
            "es_docente": es_docente,
            "docente_actual": docente_actual,
        })

    # -------------------------------
    # SI ES POST → GUARDAR
    # -------------------------------
    if request.method == "POST":
        tipo_solicitante   = request.POST.get("tipo_solicitante", "docente")
        docente_codigo     = request.POST.get("docente_codigo", "").strip()
        asignatura_nombre  = request.POST.get("asignatura_nombre", "").strip()

        fecha_str          = request.POST.get("fecha", "").strip()
        hora_inicio_str    = request.POST.get("hora_inicio", "").strip()
        hora_fin_str       = request.POST.get("hora_fin", "").strip()
        observaciones      = request.POST.get("observaciones", "").strip()

        codigos    = request.POST.getlist("codigo_herramienta")
        cantidades = request.POST.getlist("cantidad")

        docente    = None
        asignatura = None

        # ---------------- DOCENTE ----------------
        if tipo_solicitante == "docente":
            if es_docente:
                # 👉 Docente logueado: usamos siempre su propio registro
                docente = docente_actual
                if docente is None:
                    error = (
                        "Tu usuario no está vinculado a un docente activo en el sistema. "
                        "Consulta con jefatura."
                    )
            else:
                # 👉 Vista pañolero: selecciona docente desde combo
                if docente_codigo:
                    docente = Docente.objects.filter(
                        codigo=docente_codigo,
                        activo=True
                    ).first()
                    if docente is None:
                        error = "Docente no válido o inactivo."
                else:
                    error = "Debes seleccionar un docente."
        else:
            error = "Las preparaciones anticipadas solo pueden ser solicitadas por docentes."

        # ---------------- ASIGNATURA (por NOMBRE) ----------------
        if asignatura_nombre:
            asignatura, _ = Asignatura.objects.get_or_create(
                nombre=asignatura_nombre,
                defaults={"codigo": None},
            )
        else:
            if not error:
                error = "Debes indicar el nombre de la asignatura."

        # ---------------- FECHA ----------------
        try:
            fecha = timezone.datetime.strptime(fecha_str, "%Y-%m-%d").date()
        except Exception:
            fecha = None
            if not error:
                error = "Debes indicar la fecha de la clase."

        # 🔹 Regla: mínimo 2 días de anticipación para la preparación
        #if fecha and not error:
         #   hoy = timezone.localdate()
          #  min_fecha = hoy + timedelta(days=2)
           # if fecha < min_fecha:
            #    error = (
             #       f"Las preparaciones deben crearse con al menos 2 días de anticipación. "
              #      f"Fecha mínima permitida: {min_fecha.strftime('%Y-%m-%d')}."
               # )

        # ---------------- HORAS ----------------
        try:
            hora_inicio = timezone.datetime.strptime(hora_inicio_str, "%H:%M").time()
        except Exception:
            hora_inicio = None

        try:
            hora_fin = timezone.datetime.strptime(hora_fin_str, "%H:%M").time()
        except Exception:
            hora_fin = None

        if (not hora_inicio or not hora_fin) and not error:
            error = "Debes indicar hora de inicio y hora de término de la clase."

        # ---------------- LÍNEAS DE DETALLE ----------------
        lineas_validas = []
        for c, cant in zip(codigos, cantidades):
            c    = c.strip()
            cant = cant.strip()
            if not c or not cant:
                continue
            try:
                cant_int = int(cant)
            except ValueError:
                continue
            if cant_int <= 0:
                continue
            lineas_validas.append((c, cant_int))

        if not lineas_validas and not error:
            error = "Debes ingresar al menos una herramienta con cantidad mayor a 0."

        if error:
            return render(request, "inventario/crear_preparacion.html", {
                "mensaje": None,
                "error": error,
                "docentes": Docente.objects.filter(activo=True).order_by("nombre"),
                "asignaturas": Asignatura.objects.all().order_by("nombre"),
                "es_docente": es_docente,
                "docente_actual": docente_actual,
            })

        # ---------------------------------------------
        # REGISTRO DE PREPARACIÓN (SIN tocar stock_disponible)
        # ---------------------------------------------
        try:
            with transaction.atomic():
                codigo_preparacion = "C" + timezone.now().strftime("%Y%m%d%H%M%S")

                prep = Preparacion.objects.create(
                    codigo_preparacion=codigo_preparacion,
                    fecha=fecha,
                    hora_inicio=hora_inicio,
                    hora_fin=hora_fin,
                    panolero=panolero,   # 👉 aquí ya nunca será None
                    docente=docente,
                    asignatura=asignatura,
                    estado="pendiente",
                    observaciones=observaciones,
                )

                for codigo, cantidad in lineas_validas:
                    herramienta = Herramienta.objects.filter(codigo=codigo).first()
                    if herramienta is None:
                        herramienta = Herramienta.objects.filter(
                            codigo_barra=codigo
                        ).first()

                    if herramienta is None:
                        raise ValueError(
                            f"No se encontró ninguna herramienta con código/código de barra '{codigo}'."
                        )

                    # 🔹 Reservas ya existentes para ESTA herramienta,
                    #     en la MISMA fecha y MISMA hora de inicio,
                    #     solo en preparaciones PENDIENTES.
                    reservas_existentes = (
                        PreparacionDetalle.objects
                        .filter(
                            herramienta=herramienta,
                            preparacion__estado="pendiente",
                            preparacion__fecha=fecha,
                            preparacion__hora_inicio=hora_inicio,
                        )
                        .aggregate(total=Sum("cantidad_solicitada"))["total"] or 0
                    )

                    # Stock disponible real para ese bloque horario
                    disponible_para_bloque = herramienta.stock_disponible - reservas_existentes

                    if disponible_para_bloque < cantidad:
                        raise ValueError(
                            f"No hay suficiente stock disponible para {herramienta.nombre} "
                            f"en esa fecha y hora. "
                            f"Stock físico: {herramienta.stock_disponible}, "
                            f"ya reservado en otras preparaciones del mismo bloque: {reservas_existentes}, "
                            f"solicitado ahora: {cantidad}."
                        )

                    # Se registra la línea de preparación (reserva lógica)
                    PreparacionDetalle.objects.create(
                        preparacion=prep,
                        herramienta=herramienta,
                        cantidad_solicitada=cantidad,
                    )

                mensaje = (
                    f"Preparación creada correctamente. "
                    f"Código: {prep.codigo_preparacion}"
                )

        except Exception as e:
            error = f"Error al crear la preparación: {e}"

        return render(request, "inventario/crear_preparacion.html", {
            "mensaje": mensaje,
            "error": error,
            "docentes": Docente.objects.filter(activo=True).order_by("nombre"),
            "asignaturas": Asignatura.objects.all().order_by("nombre"),
            "es_docente": es_docente,
            "docente_actual": docente_actual,
        })

    # -------------------------------
    # GET → FORMULARIO VACÍO
    # -------------------------------
    return render(request, "inventario/crear_preparacion.html", {
        "mensaje": None,
        "error": None,
        "docentes": Docente.objects.filter(activo=True).order_by("nombre"),
        "asignaturas": Asignatura.objects.all().order_by("nombre"),
        "es_docente": es_docente,
        "docente_actual": docente_actual,
    })


def vista_recomendaciones(request, asignatura_id):
    asig = get_object_or_404(Asignatura, id=asignatura_id)
    recs = rec.recomendar_herramientas(asig.id, top_n=20)  # o todas
    return render(request, "inventario/recomendaciones.html", {
        "asignatura": asig,
        "recomendaciones": recs,
    })

#parte del registro preparacion 

@login_required 
def api_preparacion_por_codigo(request):
    codigo = request.GET.get("codigo", "").strip()
    if not codigo:
        return JsonResponse(
            {"ok": False, "error": "Debe indicar un código."},
            status=400
        )

    try:
        prep = Preparacion.objects.select_related(
            "docente", "asignatura", "panolero"
        ).prefetch_related(
            "detalles__herramienta"
        ).get(codigo_preparacion=codigo)
    except Preparacion.DoesNotExist:
        return JsonResponse(
            {"ok": False, "error": "Preparación no encontrada."},
            status=404
        )

    # 🔹 NUEVO: bloquear preparaciones ya usadas o anuladas
    if prep.estado in ["usado", "anulado"]:
        return JsonResponse(
            {
                "ok": False,
                "error": (
                    "Esta preparación ya fue marcada como entregada/anulada "
                    "y no puede volver a generar un préstamo."
                ),
            },
            status=400,
        )

    # Si llega aquí, la preparación es válida (pendiente)
    detalles = []
    for det in prep.detalles.all():
        h = det.herramienta
        detalles.append({
            "codigo": h.codigo,
            "nombre": h.nombre,
            "stock_disponible": h.stock_disponible,
            "cantidad": det.cantidad_solicitada,
        })

    data = {
        "ok": True,
        "codigo_preparacion": prep.codigo_preparacion,
        "fecha": prep.fecha.strftime("%Y-%m-%d") if prep.fecha else "",
        "hora_inicio": prep.hora_inicio.strftime("%H:%M") if prep.hora_inicio else "",
        "hora_fin": prep.hora_fin.strftime("%H:%M") if getattr(prep, "hora_fin", None) else "",
        "docente_codigo": prep.docente.codigo if prep.docente else "",
        "docente_nombre": prep.docente.nombre if prep.docente else "",
        "asignatura_id": prep.asignatura.id if prep.asignatura else None,
        "asignatura_nombre": prep.asignatura.nombre if prep.asignatura else "",
        "panolero_nombre": prep.panolero.nombre if prep.panolero else "",
        "detalles": detalles,
    }

    return JsonResponse(data)



#vista para observacion de la descripcion de una preparacion de pedidos para las clases 
@login_required
def detalle_preparacion(request, prep_id):
    preparacion = get_object_or_404(
        Preparacion.objects.select_related("panolero", "docente", "asignatura")
                           .prefetch_related("detalles__herramienta"),
        id=prep_id
    )

    return render(request, "inventario/detalle_preparacion.html", {
        "preparacion": preparacion,
    })

#Anular una preparacion anticipada de las clases
@login_required
def anular_preparacion(request, prep_id):
    # Traemos también los detalles con sus herramientas (por si quieres mostrarlos en una vista futura)
    preparacion = get_object_or_404(
        Preparacion.objects.prefetch_related("detalles__herramienta"),
        id=prep_id
    )

    #k
    if preparacion.estado == "usado":
        return redirect("lista_preparaciones")

    if request.method == "POST":
        preparacion.estado = "anulado"
        preparacion.save(update_fields=["estado", "updated_at"])

        return redirect("lista_preparaciones")

    # Si no quieres pantalla de confirmación, redirigimos directo
    return redirect("lista_preparaciones")

#Parete de dar de baja una heramientas 
@login_required
def registrar_baja(request):
    """
    Página independiente para dar de baja herramientas físicas.
    - No depende directamente de un préstamo.
    - Permite registrar contexto: código de préstamo, docente, asignatura,
      fecha/hora de actividad, etc.
    - Descuenta stock y stock_disponible.
    """
    panolero = obtener_panolero_desde_user(request.user)
    if panolero is None:
        return redirect("menu_principal")

    mensaje = None
    error = None

    if request.method == "POST":
        # ---- CONTEXTO DE LA BAJA ----
        codigo_prestamo_ctx = request.POST.get("codigo_prestamo_ctx", "").strip()
        docente_codigo      = request.POST.get("docente_codigo", "").strip()
        asignatura_nombre   = request.POST.get("asignatura_nombre", "").strip()
        fecha_clase_str     = request.POST.get("fecha_clase", "").strip()
        hora_inicio_str     = request.POST.get("hora_inicio_clase", "").strip()
        hora_fin_str        = request.POST.get("hora_fin_clase", "").strip()

        motivo_general = request.POST.get("motivo_general", "").strip()
        observaciones  = request.POST.get("observaciones", "").strip()

        docente    = None
        asignatura = None
        fecha_clase = None
        hora_inicio_clase = None

        # Docente opcional
        if docente_codigo:
            docente = Docente.objects.filter(
                codigo=docente_codigo,
                activo=True
            ).first()

        # Asignatura opcional, por nombre
        if asignatura_nombre:
            asignatura, _ = Asignatura.objects.get_or_create(
                nombre=asignatura_nombre,
                defaults={"codigo": None},
            )

        # Fecha clase opcional
        if fecha_clase_str:
            try:
                fecha_clase = timezone.datetime.strptime(
                    fecha_clase_str, "%Y-%m-%d"
                ).date()
            except Exception:
                fecha_clase = None

        # Hora inicio opcional
        if hora_inicio_str:
            try:
                hora_inicio_clase = timezone.datetime.strptime(
                    hora_inicio_str, "%H:%M"
                ).time()
            except Exception:
                hora_inicio_clase = None

        # Hora fin SOLO se guarda como texto en observaciones (el modelo no tiene campo)
        extra_ctx = []
        if codigo_prestamo_ctx:
            extra_ctx.append(f"Código préstamo ref: {codigo_prestamo_ctx}")
        if hora_fin_str:
            extra_ctx.append(f"Hora término actividad: {hora_fin_str}")

        if extra_ctx:
            if observaciones:
                observaciones = observaciones + " | " + " | ".join(extra_ctx)
            else:
                observaciones = " | ".join(extra_ctx)

        # ---- LÍNEAS DE BAJA ----
        codigos    = request.POST.getlist("codigo_herramienta")
        cantidades = request.POST.getlist("cantidad_baja")

        lineas_validas = []
        for c, cant in zip(codigos, cantidades):
            c    = c.strip()
            cant = cant.strip()
            if not c or not cant:
                continue
            try:
                cant_int = int(cant)
            except ValueError:
                continue
            if cant_int <= 0:
                continue
            lineas_validas.append((c, cant_int))

        if not lineas_validas:
            error = "Debes ingresar al menos una herramienta con una cantidad de baja mayor a 0."
        elif not motivo_general:
            error = "Debes indicar un motivo general para la baja."

        if not error:
            try:
                with transaction.atomic():
                    baja = Baja.objects.create(
                        fecha_registro=timezone.now().date(),
                        hora_registro=timezone.now().time(),
                        panolero=panolero,
                        docente=docente,
                        asignatura=asignatura,
                        fecha_clase=fecha_clase,
                        hora_inicio_clase=hora_inicio_clase,
                        seccion=None,
                        motivo_general=motivo_general,
                        observaciones=observaciones,
                    )

                    total_bajas = 0

                    for codigo, cantidad in lineas_validas:
                        herramienta = Herramienta.objects.filter(codigo=codigo).first()
                        if herramienta is None:
                            herramienta = Herramienta.objects.filter(
                                codigo_barra=codigo
                            ).first()

                        if herramienta is None:
                            # Si una línea no existe, la saltamos pero seguimos con las otras
                            continue

                        tipo_h = (herramienta.tipo or "").strip().lower()
                        # ❌ No permitir bajas de llaves
                        if tipo_h.startswith("llave"):
                            continue

                        # Máximo que podemos dar de baja = lo disponible en pañol
                        max_baja = min(
                            herramienta.stock,
                            herramienta.stock_disponible
                        )
                        if max_baja <= 0:
                            continue

                        cant_real = cantidad
                        if cant_real > max_baja:
                            cant_real = max_baja

                        if cant_real <= 0:
                            continue

                        BajaDetalle.objects.create(
                            baja=baja,
                            herramienta=herramienta,
                            cantidad_baja=cant_real,
                            motivo=motivo_general,
                            observacion=observaciones,
                        )

                        # Descontamos de stock total y del disponible
                        herramienta.stock -= cant_real
                        herramienta.stock_disponible -= cant_real
                        if herramienta.stock < 0:
                            herramienta.stock = 0
                        if herramienta.stock_disponible < 0:
                            herramienta.stock_disponible = 0
                        herramienta.save()

                        total_bajas += cant_real

                    if total_bajas == 0:
                        error = (
                            "No se pudo registrar ninguna baja. "
                            "Revisa que las herramientas tengan stock disponible "
                            "y que no sean llaves."
                        )
                        # Si no hubo bajas reales, borramos la cabecera
                        baja.delete()
                    else:
                        mensaje = "Bajas registradas correctamente."

            except Exception as e:
                error = f"Ocurrió un error al registrar las bajas: {e}"

    # Para mostrar un listado de herramientas de referencia (opcional)
    herramientas = Herramienta.objects.all().order_by("nombre")
    docentes = Docente.objects.filter(activo=True).order_by("nombre")

    return render(request, "inventario/registrar_baja.html", {
        "mensaje": mensaje,
        "error": error,
        "herramientas": herramientas,
        "docentes": docentes,
        "panolero": panolero,
    })


#####
#Stock ectivo considerando preparaciones 
def stock_disponible_respetando_preps(herramienta, ahora=None):
    """
    Devuelve el stock disponible efectivo de una herramienta, descontando
    las preparaciones PENDIENTES del día cuya hora_inicio está en los
    próximos 15 minutos.
    """
    if ahora is None:
        ahora = timezone.localtime()

    hoy = ahora.date()
    ventana_fin = ahora + timedelta(minutes=15)
  


    # Stock disponible físico en pañol
    base = herramienta.stock_disponible

    # Cantidad reservada en preparaciones pendientes del día,
    # con hora_inicio entre ahora y ahora+15 min
    reservas = (
        PreparacionDetalle.objects
        .filter(
            herramienta=herramienta,
            preparacion__estado="pendiente",
            preparacion__fecha=hoy,
            preparacion__hora_inicio__gte=ahora.time(),
            preparacion__hora_inicio__lte=ventana_fin.time(),
        )
        .aggregate(total=Sum("cantidad_solicitada"))["total"] or 0
    )

    return base - reservas


#Lista de bajas 
@login_required
def lista_bajas(request):
    """
    Lista de bajas registradas, ordenadas de más reciente a más antigua.
    Permite filtrar por texto y por rango de fechas (fecha_registro).
    """
    q = request.GET.get("q", "").strip()
    fecha_desde = request.GET.get("fecha_desde", "").strip()
    fecha_hasta = request.GET.get("fecha_hasta", "").strip()

    bajas = Baja.objects.select_related(
        "panolero", "docente", "asignatura"
    ).all()

    # --- Filtro por fechas (usando fecha_registro) ---
    if fecha_desde:
        try:
            fd = timezone.datetime.strptime(fecha_desde, "%Y-%m-%d").date()
            bajas = bajas.filter(fecha_registro__gte=fd)
        except ValueError:
            pass

    if fecha_hasta:
        try:
            fh = timezone.datetime.strptime(fecha_hasta, "%Y-%m-%d").date()
            bajas = bajas.filter(fecha_registro__lte=fh)
        except ValueError:
            pass

    # --- Filtro por texto ---
    if q:
        bajas = bajas.filter(
            Q(motivo_general__icontains=q)
            | Q(observaciones__icontains=q)
            | Q(panolero__nombre__icontains=q)
            | Q(docente__nombre__icontains=q)
            | Q(asignatura__nombre__icontains=q)
        )

    bajas = bajas.order_by("-fecha_registro", "-id")

    return render(request, "inventario/lista_bajas.html", {
        "bajas": bajas,
        "query": q,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
    })

#Api codigo prestamo para 
@login_required
def api_prestamo_por_codigo(request):
    """
    Devuelve información básica de un préstamo dado su código,
    para autocompletar el formulario de bajas y mostrar herramientas prestadas.
    """
    codigo = request.GET.get("codigo", "").strip()
    if not codigo:
        return JsonResponse(
            {"ok": False, "error": "Debe indicar un código de préstamo."},
            status=400
        )

    try:
        p = Prestamo.objects.select_related(
            "docente",
            "estudiante",
            "asignatura",
            "panolero",
        ).prefetch_related(
            "detalles__herramienta"       # 👈 IMPORTANTE: cargamos los detalles
        ).get(codigo_prestamo=codigo)
    except Prestamo.DoesNotExist:
        return JsonResponse(
            {"ok": False, "error": "Préstamo no encontrado."},
            status=404
        )

    # Armamos la lista de herramientas prestadas
    detalles = []
    for det in p.detalles.all():
        detalles.append({
            "id": det.id,
            "codigo": det.herramienta.codigo,
            "nombre": det.herramienta.nombre,
            "cantidad_entregada": det.cantidad_entregada,
            "cantidad_devuelta": det.cantidad_devuelta,
        })

    data = {
        "ok": True,
        "codigo_prestamo": p.codigo_prestamo,
        "fecha": p.fecha.strftime("%Y-%m-%d") if p.fecha else "",
        "hora_inicio": p.hora_inicio.strftime("%H:%M") if p.hora_inicio else "",
        "hora_fin": p.hora_fin.strftime("%H:%M") if p.hora_fin else "",

        "docente_codigo": p.docente.codigo if p.docente else "",
        "docente_nombre": p.docente.nombre if p.docente else "",

        "estudiante_rut": p.estudiante.rut if p.estudiante else "",
        "estudiante_nombre": p.estudiante.nombre if p.estudiante else "",
        "estudiante_carrera": (
            p.estudiante.carrera if (p.estudiante and p.estudiante.carrera) else ""
        ),

        "asignatura_nombre": p.asignatura.nombre if p.asignatura else "",
        "panolero_nombre": p.panolero.nombre if p.panolero else "",

        # 👇  herramientas del préstamo
        "detalles": detalles,
    }

    return JsonResponse(data)

#Ver detalle de la baja 
@login_required
def detalle_baja(request, baja_id):
    """
    Muestra el detalle completo de una baja:
    - Datos de cabecera (pañolero, docente, asignatura, fecha, observaciones)
    - Listado de herramientas dadas de baja.
    """
    baja = get_object_or_404(
        Baja.objects.select_related("panolero", "docente", "asignatura"),
        id=baja_id
    )

    # Si el related_name de BajaDetalle no está definido como "detalles",
    # usamos filter por seguridad
    detalles = BajaDetalle.objects.select_related("herramienta").filter(baja=baja)

    return render(request, "inventario/detalle_baja.html", {
        "baja": baja,
        "detalles": detalles,
    })


#-----------------------------------
#INFORMES
#-----------------------------------


@login_required
def informe_prestamos(request):
    """
    Informe completo de préstamos:
    - Filtros por fecha, estado y texto.
    - Resumen general (totales).
    - Top 5 herramientas más prestadas.
    - Top 5 llaves (incluye autos si las marcaste como llave/llave_auto).
    - Top 5 autos (llaves_auto).
    - Top 5 docentes con más préstamos.
    - Top 5 asignaturas.
    - Top 5 pañoleros por cantidad de préstamos gestionados.
    """

    fecha_desde = request.GET.get("fecha_desde", "").strip()
    fecha_hasta = request.GET.get("fecha_hasta", "").strip()
    estado      = request.GET.get("estado", "").strip()
    q           = request.GET.get("q", "").strip()

    # ------------------ BASE QUERY DE PRÉSTAMOS ------------------
    prestamos = Prestamo.objects.select_related(
        "panolero", "docente", "estudiante", "asignatura"
    ).all()

    # Filtro por fecha desde/hasta
    if fecha_desde:
        try:
            fd = timezone.datetime.strptime(fecha_desde, "%Y-%m-%d").date()
            prestamos = prestamos.filter(fecha__gte=fd)
        except ValueError:
            pass

    if fecha_hasta:
        try:
            fh = timezone.datetime.strptime(fecha_hasta, "%Y-%m-%d").date()
            prestamos = prestamos.filter(fecha__lte=fh)
        except ValueError:
            pass

    # --------- Filtro por estado ---------
    if estado:
        prestamos = prestamos.filter(estado__iexact=estado)

    # Búsqueda por texto (docente/estudiante/asignatura/pañolero)
    if q:
        prestamos = prestamos.filter(
            Q(docente__nombre__icontains=q)
            | Q(estudiante__nombre__icontains=q)
            | Q(estudiante__carrera__icontains=q)
            | Q(asignatura__nombre__icontains=q)
            | Q(panolero__nombre__icontains=q)
        )

    prestamos = prestamos.order_by("-fecha", "-id")

    # ------------------ RESUMEN GENERALES ------------------
    total_prestamos    = prestamos.count()
    total_docentes     = (
        prestamos.exclude(docente__isnull=True)
        .values("docente_id")
        .distinct()
        .count()
    )
    total_estudiantes  = (
        prestamos.exclude(estudiante__isnull=True)
        .values("estudiante_id")
        .distinct()
        .count()
    )
    total_asignaturas  = (
        prestamos.exclude(asignatura__isnull=True)
        .values("asignatura_id")
        .distinct()
        .count()
    )
    total_panoleros    = (
        prestamos.exclude(panolero__isnull=True)
        .values("panolero_id")
        .distinct()
        .count()
    )

    # Detalles filtrados por los mismos préstamos (para top herramientas, llaves, etc.)
    detalles = PrestamoDetalle.objects.filter(
        prestamo__in=prestamos
    ).select_related("herramienta")

    # ------------------ TOP 5 HERRAMIENTAS MÁS PRESTADAS ------------------
    top_herramientas = (
        detalles.values("herramienta__codigo", "herramienta__nombre", "herramienta__tipo")
        .annotate(
            total_prestada=Sum("cantidad_entregada"),
            num_prestamos=Count("prestamo", distinct=True),
        )
        .order_by("-total_prestada")[:5]
    )

    # ------------------ TOP 5 LLAVES (incluye todo tipo 'llave*') ------------------
    top_llaves = (
        detalles.filter(herramienta__tipo__icontains="llave")
        .values("herramienta__codigo", "herramienta__nombre", "herramienta__tipo")
        .annotate(
            total_prestada=Sum("cantidad_entregada"),
            num_prestamos=Count("prestamo", distinct=True),
        )
        .order_by("-num_prestamos")[:5]
    )

    # ------------------ TOP 5 AUTOS (llaves de auto, tipo contiene 'llave_auto') ------------------
    top_autos = (
        detalles.filter(herramienta__tipo__icontains="llave_auto")
        .values("herramienta__codigo", "herramienta__nombre")
        .annotate(
            total_prestada=Sum("cantidad_entregada"),
            num_prestamos=Count("prestamo", distinct=True),
        )
        .order_by("-num_prestamos")[:5]
    )

    # ------------------ TOP 5 DOCENTES ------------------
    top_docentes = (
        prestamos.filter(docente__isnull=False)
        .values("docente__codigo", "docente__nombre")
        .annotate(num_prestamos=Count("id"))
        .order_by("-num_prestamos")[:5]
    )

    # ------------------ TOP 5 ASIGNATURAS ------------------
    top_asignaturas = (
        prestamos.filter(asignatura__isnull=False)
        .values("asignatura__id", "asignatura__nombre")
        .annotate(num_prestamos=Count("id"))
        .order_by("-num_prestamos")[:5]
    )

    # ------------------ TOP 5 PAÑOLEROS ------------------
    top_panoleros = (
        prestamos.filter(panolero__isnull=False)
        .values("panolero__id", "panolero__nombre")
        .annotate(num_prestamos=Count("id"))
        .order_by("-num_prestamos")[:5]
    )

    # 🔹 HERRAMIENTAS SIN STOCK DISPONIBLE
    # Versión básica: todas las herramientas del pañol con stock_disponible <= 0
    herramientas_sin_stock = (
        Herramienta.objects
        .filter(stock_disponible__lte=0)
        .order_by("nombre")
    )
    # Si quieres mostrar solo las que hayan participado en los préstamos filtrados:
    # ids_usadas = detalles.values_list("herramienta_id", flat=True).distinct()
    # herramientas_sin_stock = Herramienta.objects.filter(
    #     id__in=ids_usadas,
    #     stock_disponible__lte=0
    # ).order_by("nombre")

    context = {
        "prestamos": prestamos,

        "resumen": {
            "total_prestamos": total_prestamos,
            "total_docentes": total_docentes,
            "total_estudiantes": total_estudiantes,
            "total_asignaturas": total_asignaturas,
            "total_panoleros": total_panoleros,
        },

        "top_herramientas": top_herramientas,
        "top_llaves": top_llaves,
        "top_autos": top_autos,
        "top_docentes": top_docentes,
        "top_asignaturas": top_asignaturas,
        "top_panoleros": top_panoleros,

        # 👇 NUEVO: lista de herramientas sin stock
        "herramientas_sin_stock": herramientas_sin_stock,

        # para que el template pueda marcar el estado seleccionado en el <select>
        "estado_sel": estado,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "q": q,
    }

    return render(request, "inventario/informe_prestamos.html", context)


#KPI DASBOAR

@login_required
def panel_kpis(request):
    semestre = request.GET.get("semestre", "")              # ej. "2025-1"
    carrera_filtro = request.GET.get("carrera", "")         # carrera de estudiante
    asignatura_filtro = request.GET.get("asignatura", "")   # nombre asignatura

    # Rango de fechas (YYYY-MM-DD) desde el formulario
    fecha_desde_str = request.GET.get("fecha_desde", "").strip()
    fecha_hasta_str = request.GET.get("fecha_hasta", "").strip()

    # 1) Si NO hay ningún filtro, aplicamos por defecto "últimos 90 días"
    if not semestre and not carrera_filtro and not asignatura_filtro and not fecha_desde_str and not fecha_hasta_str:
        hoy = timezone.localdate()
        fecha_desde_default = hoy - timedelta(days=90)
        fecha_desde_str = fecha_desde_default.strftime("%Y-%m-%d")
        fecha_hasta_str = hoy.strftime("%Y-%m-%d")

    # Base de préstamos (con relaciones)
    prestamos = Prestamo.objects.select_related(
        "docente", "estudiante", "asignatura", "panolero"
    )

    # --- FILTRO POR SEMESTRE ---
    if semestre:
        try:
            año_str, sem_str = semestre.split("-")  # "2025-1"
            año = int(año_str)
            if sem_str == "1":   # 1er semestre (ejemplo marzo–julio)
                prestamos = prestamos.filter(
                    fecha__year=año,
                    fecha__month__in=[3, 4, 5, 6, 7],
                )
            elif sem_str == "2": # 2º semestre (agosto–diciembre)
                prestamos = prestamos.filter(
                    fecha__year=año,
                    fecha__month__in=[8, 9, 10, 11, 12],
                )
        except ValueError:
            pass

    # --- FILTRO POR RANGO DE FECHAS ---
    if fecha_desde_str:
        try:
            fecha_desde = timezone.datetime.strptime(
                fecha_desde_str, "%Y-%m-%d"
            ).date()
            prestamos = prestamos.filter(fecha__gte=fecha_desde)
        except ValueError:
            fecha_desde = None
    else:
        fecha_desde = None

    if fecha_hasta_str:
        try:
            fecha_hasta = timezone.datetime.strptime(
                fecha_hasta_str, "%Y-%m-%d"
            ).date()
            prestamos = prestamos.filter(fecha__lte=fecha_hasta)
        except ValueError:
            fecha_hasta = None
    else:
        fecha_hasta = None

    # --- FILTRO POR CARRERA (solo préstamos de estudiantes) ---
    if carrera_filtro:
        prestamos = prestamos.filter(estudiante__carrera=carrera_filtro)

    # --- FILTRO POR ASIGNATURA ---
    if asignatura_filtro:
        prestamos = prestamos.filter(asignatura__nombre=asignatura_filtro)

    # ========== IMPORTANTE ==========
    # Usaremos `prestamos_qs` para KPIs y rankings (TODOS los registros filtrados)
    # y luego paginaremos SOLO para la tabla.
    # =================================
    prestamos_qs = prestamos

    # ---------------- KPIs GENERALES ----------------
    total_prestamos = prestamos_qs.count()

    total_herramientas = (
        PrestamoDetalle.objects
        .filter(prestamo__in=prestamos_qs)
        .aggregate(total=Sum("cantidad_entregada"))["total"] or 0
    )

    total_prest_docente = prestamos_qs.filter(docente__isnull=False).count()
    total_prest_estudiante = prestamos_qs.filter(estudiante__isnull=False).count()
    total_prest_otros = total_prestamos - total_prest_docente - total_prest_estudiante

    # 🔹 cuántos pañoleros distintos han intervenido en estos préstamos
    total_panoleros = (
        prestamos_qs.exclude(panolero__isnull=True)
        .values("panolero")
        .distinct()
        .count()
    )

    # ---------------- TOPs PARA GRÁFICOS ----------------

    # Top 5 docentes por cantidad de préstamos
    top_docentes = (
        prestamos_qs.filter(docente__isnull=False)
        .values("docente__nombre", "docente__codigo")
        .annotate(total_prestamos=Count("id"))
        .order_by("-total_prestamos")[:5]
    )

    # Top 5 carreras (préstamos de estudiantes)
    top_carreras = (
        prestamos_qs.filter(estudiante__isnull=False, estudiante__carrera__isnull=False)
        .values("estudiante__carrera")
        .annotate(total_prestamos=Count("id"))
        .order_by("-total_prestamos")[:5]
    )

    # Detalles filtrados (para herramientas / autos)
    detalles_filtrados = PrestamoDetalle.objects.filter(prestamo__in=prestamos_qs)

    # Top 5 herramientas más despachadas (general)
    top_herramientas = (
        detalles_filtrados
        .values("herramienta__nombre", "herramienta__codigo")
        .annotate(total_cant=Sum("cantidad_entregada"))
        .order_by("-total_cant")[:5]
    )

    # 🔹 Top 5 herramientas FIJAS
    top_herramientas_fijas = (
        detalles_filtrados
        .filter(herramienta__tipo__icontains="fijo")
        .values("herramienta__nombre", "herramienta__codigo")
        .annotate(total_cant=Sum("cantidad_entregada"))
        .order_by("-total_cant")[:5]
    )

    # 🔹 Top 5 herramientas CONSUMIBLES
    top_herramientas_consumibles = (
        detalles_filtrados
        .filter(herramienta__tipo__icontains="consumible")
        .values("herramienta__nombre", "herramienta__codigo")
        .annotate(total_cant=Sum("cantidad_entregada"))
        .order_by("-total_cant")[:5]
    )

    # Top 5 llaves de auto (asumiendo tipo contiene "llave_auto")
    top_autos = (
        detalles_filtrados
        .filter(herramienta__tipo__icontains="llave_auto")
        .values("herramienta__nombre", "herramienta__codigo")
        .annotate(total_prestamos=Count("prestamo", distinct=True))
        .order_by("-total_prestamos")[:5]
    )

    # Top 5 asignaturas
    top_asignaturas = (
        prestamos_qs.filter(asignatura__isnull=False)
        .values("asignatura__nombre")
        .annotate(total_prestamos=Count("id"))
        .order_by("-total_prestamos")[:5]
    )

    # Top 5 pañoleros por cantidad de préstamos registrados
    top_panoleros = (
        prestamos_qs.filter(panolero__isnull=False)
        .values("panolero__nombre")
        .annotate(total_prestamos=Count("id"))
        .order_by("-total_prestamos")[:5]
    )

    # ---------------- PAGINACIÓN PARA LA TABLA ----------------
    # Ordenamos y paginamos SOLO para la tabla de "Listado de préstamos"
    prestamos_qs = prestamos_qs.order_by("-fecha", "-id")

    paginator = Paginator(prestamos_qs, 100)  # 100 préstamos por página (ajusta a gusto)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Para construir los links de paginación sin duplicar el parámetro "page"
    query_params = request.GET.copy()
    if "page" in query_params:
        query_params.pop("page")
    querystring = query_params.urlencode()

    # ---------------- LISTAS PARA LOS SELECT ----------------
    lista_semestres = ["2024-1", "2024-2", "2025-1", "2025-2"]

    lista_carreras = (
        Estudiante.objects
        .exclude(carrera__isnull=True)
        .exclude(carrera="")
        .values_list("carrera", flat=True)
        .distinct()
        .order_by("carrera")
    )

    lista_asignaturas = (
        Asignatura.objects
        .values_list("nombre", flat=True)
        .distinct()
        .order_by("nombre")
    )

    context = {
        # PRESTAMOS PAGINADOS (para la tabla)
        "page_obj": page_obj,
        "querystring": querystring,

        # KPIs
        "total_prestamos": total_prestamos,
        "total_herramientas": total_herramientas,
        "total_prest_docente": total_prest_docente,
        "total_prest_estudiante": total_prest_estudiante,
        "total_prest_otros": total_prest_otros,

        "total_panoleros": total_panoleros,
        "top_panoleros": top_panoleros,

        # TOPs
        "top_docentes": top_docentes,
        "top_carreras": top_carreras,
        "top_herramientas": top_herramientas,
        "top_herramientas_fijas": top_herramientas_fijas,
        "top_herramientas_consumibles": top_herramientas_consumibles,
        "top_autos": top_autos,
        "top_asignaturas": top_asignaturas,

        # Filtros para los select
        "lista_semestres": lista_semestres,
        "lista_carreras": lista_carreras,
        "lista_asignaturas": lista_asignaturas,
        "semestre_sel": semestre,
        "carrera_sel": carrera_filtro,
        "asignatura_sel": asignatura_filtro,

        # Para mantener el rango de fechas en los inputs y en exportación
        "fecha_desde": fecha_desde_str,
        "fecha_hasta": fecha_hasta_str,
    }
    return render(request, "inventario/panel_kpis.html", context)

# Exportar
@login_required
def exportar_panel_kpis(request):
    semestre = request.GET.get("semestre", "")
    carrera = request.GET.get("carrera", "")
    asignatura_nombre = request.GET.get("asignatura", "")
    fecha_desde = request.GET.get("fecha_desde", "")
    fecha_hasta = request.GET.get("fecha_hasta", "")

    # ================== BASE QUERY ==================
    prestamos = Prestamo.objects.select_related(
        "docente", "estudiante", "asignatura", "panolero"
    )

    # ----- Filtro por semestre -----
    if semestre:
        try:
            año_str, sem_str = semestre.split("-")  # "2025-1"
            año = int(año_str)
            if sem_str == "1":   # 1er semestre
                prestamos = prestamos.filter(
                    fecha__year=año,
                    fecha__month__in=[3, 4, 5, 6, 7],
                )
            elif sem_str == "2":  # 2º semestre
                prestamos = prestamos.filter(
                    fecha__year=año,
                    fecha__month__in=[8, 9, 10, 11, 12],
                )
        except ValueError:
            pass

    # ----- Filtro por carrera -----
    if carrera:
        prestamos = prestamos.filter(estudiante__carrera=carrera)

    # ----- Filtro por asignatura -----
    if asignatura_nombre:
        prestamos = prestamos.filter(asignatura__nombre=asignatura_nombre)

    # ----- Filtro por rango de fechas -----
    # Si tu input date está en formato YYYY-MM-DD (lo normal del navegador),
    # con este formato basta. Si tuvieras DD-MM-YYYY, se intenta también.
    if fecha_desde:
        parseado = None
        for fmt in ("%Y-%m-%d", "%d-%m-%Y"):
            try:
                parseado = datetime.strptime(fecha_desde, fmt).date()
                break
            except ValueError:
                continue
        if parseado:
            prestamos = prestamos.filter(fecha__gte=parseado)

    if fecha_hasta:
        parseado = None
        for fmt in ("%Y-%m-%d", "%d-%m-%Y"):
            try:
                parseado = datetime.strptime(fecha_hasta, fmt).date()
                break
            except ValueError:
                continue
        if parseado:
            prestamos = prestamos.filter(fecha__lte=parseado)

    # ================== CÁLCULO DE KPIs (SOLO CON FILTROS APLICADOS) ==================
    total_prestamos = prestamos.count()

    total_herramientas = (
        PrestamoDetalle.objects
        .filter(prestamo__in=prestamos)
        .aggregate(total=Sum("cantidad_entregada"))["total"] or 0
    )

    total_prest_docente = prestamos.filter(docente__isnull=False).count()
    total_prest_estudiante = prestamos.filter(estudiante__isnull=False).count()
    total_prest_otros = total_prestamos - total_prest_docente - total_prest_estudiante

    total_panoleros = (
        prestamos.exclude(panolero__isnull=True)
        .values("panolero")
        .distinct()
        .count()
    )

    # ----- Top 5 docentes -----
    top_docentes = (
        prestamos.filter(docente__isnull=False)
        .values("docente__nombre", "docente__codigo")
        .annotate(total_prestamos=Count("id"))
        .order_by("-total_prestamos")[:5]
    )

    # ----- Top 5 carreras -----
    top_carreras = (
        prestamos.filter(estudiante__isnull=False, estudiante__carrera__isnull=False)
        .values("estudiante__carrera")
        .annotate(total_prestamos=Count("id"))
        .order_by("-total_prestamos")[:5]
    )

    # Detalles filtrados (para herramientas / autos)
    detalles_filtrados = PrestamoDetalle.objects.filter(prestamo__in=prestamos)

    # ----- Top 5 herramientas -----
    top_herramientas = (
        detalles_filtrados
        .values("herramienta__nombre", "herramienta__codigo")
        .annotate(total_cant=Sum("cantidad_entregada"))
        .order_by("-total_cant")[:5]
    )

    # ----- Top 5 autos (llaves auto) -----
    top_autos = (
        detalles_filtrados
        .filter(herramienta__tipo__icontains="llave_auto")
        .values("herramienta__nombre", "herramienta__codigo")
        .annotate(total_prestamos=Count("prestamo", distinct=True))
        .order_by("-total_prestamos")[:5]
    )

    # ----- Top 5 asignaturas -----
    top_asignaturas = (
        prestamos.filter(asignatura__isnull=False)
        .values("asignatura__nombre")
        .annotate(total_prestamos=Count("id"))
        .order_by("-total_prestamos")[:5]
    )

    # ----- Top 5 pañoleros -----
    top_panoleros = (
        prestamos.filter(panolero__isnull=False)
        .values("panolero__nombre")
        .annotate(total_prestamos=Count("id"))
        .order_by("-total_prestamos")[:5]
    )

    formato = request.GET.get("formato", "excel").lower()

    # ======================================================
    #   EXPORTAR A EXCEL (detalle completo filtrado)
    # ======================================================
    if formato == "excel":
        import openpyxl

        wb = openpyxl.Workbook()

        # --- Hoja 1: KPIs ---
        ws_kpi = wb.active
        ws_kpi.title = "KPIs"

        ws_kpi.append(["KPI", "Valor"])
        ws_kpi.append(["Total de préstamos", total_prestamos])
        ws_kpi.append(["Herramientas entregadas", total_herramientas])
        ws_kpi.append(["Préstamos a docentes", total_prest_docente])
        ws_kpi.append(["Préstamos a estudiantes", total_prest_estudiante])
        ws_kpi.append(["Préstamos otros", total_prest_otros])
        ws_kpi.append(["Pañoleros activos", total_panoleros])

        ws_kpi.append([])
        ws_kpi.append(["Top 5 docentes", "Préstamos"])
        for d in top_docentes:
            ws_kpi.append([d["docente__nombre"], d["total_prestamos"]])

        ws_kpi.append([])
        ws_kpi.append(["Top 5 carreras", "Préstamos"])
        for c in top_carreras:
            ws_kpi.append([c["estudiante__carrera"], c["total_prestamos"]])

        ws_kpi.append([])
        ws_kpi.append(["Top 5 herramientas", "Cantidad entregada"])
        for h in top_herramientas:
            ws_kpi.append([h["herramienta__nombre"], h["total_cant"]])

        ws_kpi.append([])
        ws_kpi.append(["Top 5 asignaturas", "Préstamos"])
        for a in top_asignaturas:
            ws_kpi.append([a["asignatura__nombre"], a["total_prestamos"]])

        ws_kpi.append([])
        ws_kpi.append(["Top 5 llaves de auto", "Préstamos"])
        for au in top_autos:
            ws_kpi.append([au["herramienta__nombre"], au["total_prestamos"]])

        ws_kpi.append([])
        ws_kpi.append(["Top 5 pañoleros", "Préstamos"])
        for p in top_panoleros:
            ws_kpi.append([p["panolero__nombre"], p["total_prestamos"]])

        # --- Hoja 2: detalle de préstamos filtrados ---
        ws = wb.create_sheet(title="Préstamos")

        encabezados = [
            "Código", "Fecha", "Hora inicio", "Hora fin",
            "Solicitante", "Asignatura", "Pañolero", "Estado",
        ]
        ws.append(encabezados)

        for p in prestamos:
            if p.docente:
                solicitante = f"{p.docente.nombre} (Docente)"
            elif p.estudiante:
                solicitante = f"{p.estudiante.nombre} ({p.estudiante.carrera})"
            else:
                solicitante = "-"

            asignatura = p.asignatura.nombre if p.asignatura else "-"
            panolero = p.panolero.nombre if p.panolero else "-"

            ws.append([
                p.codigo_prestamo,
                p.fecha.strftime("%Y-%m-%d") if p.fecha else "",
                p.hora_inicio.strftime("%H:%M") if p.hora_inicio else "",
                p.hora_fin.strftime("%H:%M") if p.hora_fin else "",
                solicitante,
                asignatura,
                panolero,
                p.estado,
            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename=\"panel_kpis.xlsx\"'
        wb.save(response)
        return response

    # ======================================================
    #   EXPORTAR A PDF (KPIs + rankings, todos con filtros)
    # ======================================================
    if formato == "pdf":
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.utils import ImageReader
        import urllib.request

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename=\"panel_kpis.pdf\"'

        c = canvas.Canvas(response, pagesize=letter)
        width, height = letter

        # -------- LOGO --------
        try:
            logo_url = (
                "https://afeva.cl/wp-content/uploads/"
                "bfi_thumb/logo-02-3de0hedts90kzsj5so5p6cawzz6fhdkja52f67wjd5s94pl8w.png"
            )
            logo_stream = urllib.request.urlopen(logo_url)
            logo_image = ImageReader(logo_stream)

            logo_width = 80
            logo_height = 40
            x = width - logo_width - 40
            y = height - logo_height - 30

            c.drawImage(
                logo_image,
                x,
                y,
                width=logo_width,
                height=logo_height,
                preserveAspectRatio=True,
                mask="auto",
            )
        except Exception as e:
            print("No se pudo cargar el logo desde URL:", e)

        # -------- TÍTULO Y FILTROS --------
        y = height - 60
        c.setFont("Helvetica-Bold", 14)
        c.drawString(40, y, "Panel de KPIs - Resumen")
        y -= 20

        c.setFont("Helvetica", 9)
        filtros = [
            f"Semestre: {semestre or 'Todos'}",
            f"Carrera: {carrera or 'Todas'}",
            f"Asignatura: {asignatura_nombre or 'Todas'}",
            f"Fecha desde: {fecha_desde or '---'}",
            f"Fecha hasta: {fecha_hasta or '---'}",
        ]
        for linea in filtros:
            c.drawString(40, y, linea)
            y -= 12

        # -------- KPIs GENERALES (FILTRADOS) --------
        y -= 8
        c.setFont("Helvetica-Bold", 11)
        c.drawString(40, y, "KPIs generales")
        y -= 15
        c.setFont("Helvetica", 9)

        kpis = [
            f"Total de préstamos: {total_prestamos}",
            f"Herramientas entregadas: {total_herramientas}",
            f"Préstamos a docentes: {total_prest_docente}",
            f"Préstamos a estudiantes: {total_prest_estudiante}",
            f"Préstamos otros: {total_prest_otros}",
            f"Pañoleros activos: {total_panoleros}",
        ]
        for l in kpis:
            c.drawString(40, y, l)
            y -= 12

        def check_page(y_actual):
            if y_actual < 60:
                c.showPage()
                return height - 40
            return y_actual

        # -------- RANKINGS (FILTRADOS) --------
        secciones_rank = [
            ("Top 5 docentes (por préstamos)", top_docentes,
             lambda d: f"{d['docente__nombre']}: {d['total_prestamos']}"),
            ("Top 5 carreras (préstamos a estudiantes)", top_carreras,
             lambda r: f"{r['estudiante__carrera']}: {r['total_prestamos']}"),
            ("Top 5 herramientas (cantidad entregada)", top_herramientas,
             lambda h: f"{h['herramienta__nombre']}: {h['total_cant']}"),
            ("Top 5 asignaturas (préstamos)", top_asignaturas,
             lambda a: f"{a['asignatura__nombre']}: {a['total_prestamos']}"),
            ("Top 5 llaves de auto (préstamos)", top_autos,
             lambda au: f"{au['herramienta__nombre']}: {au['total_prestamos']}"),
            ("Top 5 pañoleros (préstamos registrados)", top_panoleros,
             lambda p: f"{p['panolero__nombre']}: {p['total_prestamos']}"),
        ]

        for titulo, lista, fmt in secciones_rank:
            y = check_page(y - 10)
            c.setFont("Helvetica-Bold", 11)
            c.drawString(40, y, titulo)
            y -= 15
            c.setFont("Helvetica", 9)

            if not lista:
                c.drawString(50, y, "Sin registros para este ranking.")
                y -= 12
            else:
                for item in lista:
                    y = check_page(y)
                    c.drawString(50, y, f"- {fmt(item)}")
                    y -= 12

        c.showPage()
        c.save()
        return response

    return HttpResponse("Formato no soportado", status=400)

#IA
@login_required
def asignatura_recomendaciones(request, asig_id):
    """
    Endpoint JSON para recomendaciones de herramientas de una asignatura.
    Estructura de salida:
    {
        "id": 12,
        "asignatura": "Mecánica I",
        "recomendaciones": [
            {
                "herramienta_id": 34,
                "codigo": "H00034",
                "nombre": "ALICATE DE PUNTA",
                "score": 35.0
            },
            ...
        ]
    }
    """
    asignatura = get_object_or_404(Asignatura, id=asig_id)

    # Recomendaciones del modelo (tal como lo tienes)
    recs = rec.recomendar_herramientas(asig_id, top_n=50) or []

    # Tomamos los IDs de herramienta que vienen en r["herramienta_id"]
    ids_herr = [
        r["herramienta_id"]
        for r in recs
        if r.get("herramienta_id") is not None
    ]

    # Cargamos las herramientas desde BD y las indexamos por id
    herramientas = {
        h.id: h
        for h in Herramienta.objects.filter(id__in=ids_herr)
    }

    recomendaciones = []
    for r_item in recs:
        h = herramientas.get(r_item.get("herramienta_id"))
        recomendaciones.append({
            "herramienta_id": r_item.get("herramienta_id"),
            # código real de la herramienta (campo Herramienta.codigo)
            "codigo": h.codigo if h else "",
            # nombre desde BD si existe, sino el que venga del modelo
            "nombre": h.nombre if h else r_item.get("nombre", ""),
            "score": float(r_item.get("score", 0.0)),
        })

    data = {
        "id": asignatura.id,
        "asignatura": asignatura.nombre,
        "recomendaciones": recomendaciones,
    }
    return JsonResponse(data)